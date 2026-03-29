# UPDATED BY CHATGPT: stable photo-cache + final export columns + better categorization
import io
import re
import sqlite3
import hashlib
from pathlib import Path

import boto3
from botocore.exceptions import ClientError
import pandas as pd
import requests
import streamlit as st
from rapidfuzz import fuzz
from openpyxl import load_workbook

st.set_page_config(page_title="Price Aggregator", layout="wide")

SUPPLIERS = {
    "s1": {"label": "1. Velozapchasti", "sheet_name": "Sheet1", "header_row": 7, "source_type": "file", "default_price_tier": "price", "allowed_price_tiers": ["price"]},
    "s2": {"label": "2. Форвард СПб", "sheet_name": "Лист_1", "header_row": 5, "source_type": "file", "default_price_tier": "price", "allowed_price_tiers": ["price", "price_opt2", "price_opt3", "price_rrc"]},
    "s3": {"label": "3. Колхозник / Монстр", "sheet_name": "Лист1", "header_row": 1, "source_type": "url", "default_price_tier": "price", "allowed_price_tiers": ["price", "price_opt10", "price_opt50"]},
    "s4": {"label": "4. BRATAN", "sheet_name": "Лена 25.05.24", "header_row": 0, "source_type": "url", "default_price_tier": "price", "allowed_price_tiers": ["price", "own_price", "price_rrc"]},
}

PRICE_TIER_LABELS = {
    "price": "Цена", "price_opt2": "Опт 2", "price_opt3": "Опт 3",
    "price_opt10": "От 10 шт", "price_opt50": "От 50 шт", "own_price": "Своим", "price_rrc": "РРЦ",
}

CATEGORY_RULES = [
    ("Аккумуляторы", ["аккумулятор", "акб", "battery", "lifepo4", "li-ion", "li ion", "литиев", "батарея"]),
    ("Зарядные устройства", ["зарядн", "зарядка", "charger", "адаптер питания"]),
    ("Моторы", ["мотор-колес", "мотор колес", "мотор-втулк", "hub motor", "mid drive", "редукторный мотор", "мотор"]),
    ("Контроллеры", ["контроллер", "controller", "синусный", "wave controller"]),
    ("Дисплеи", ["дисплей", "display", "lcd", "led", "панель управления"]),
    ("Кабели и разъёмы", ["кабель", "провод", "разъём", "разъем", "коннектор", "connector", "wire", "cable", "удлинитель"]),
    ("Велосипеды", ["велосипед", "электровелосипед", "байк", "bmx", "bike", "e-bike", "ebike"]),
    ("Колёса", ["колесо", "wheelset", "wheel"]),
    ("Покрышки", ["покрыш", "шина", "tire", "tyre"]),
    ("Камеры", ["камера", "tube", "innertube"]),
    ("Обода и спицы", ["обод", "спиц", "rim", "spoke"]),
    ("Втулки", ["втулк", "hub"]),
    ("Тормоза", ["тормоз", "brake", "гидравлик", "механический тормоз"]),
    ("Тормозные колодки", ["колодк", "brake pad", "pad set"]),
    ("Роторы", ["ротор", "диск тормоз", "rotor", "disc brake rotor"]),
    ("Цепи", ["цепь", "chain"]),
    ("Кассеты и трещотки", ["кассет", "трещот", "cassette", "freewheel"]),
    ("Звёзды", ["звезд", "sprocket", "chainring", "передняя звезд"]),
    ("Шатуны и каретки", ["шатун", "каретк", "crank", "bottom bracket"]),
    ("Педали", ["педал", "pedal"]),
    ("Переключатели", ["переключател", "derailleur"]),
    ("Манетки", ["манетк", "gripshift", "trigger shifter", "шифтер"]),
    ("Рули", ["руль", "handlebar", "guidon"]),
    ("Выносы", ["вынос", "stem"]),
    ("Грипсы", ["грипс", "grip", "ручк руля"]),
    ("Седла", ["седл", "saddle"]),
    ("Подседельные штыри", ["подседел", "seatpost", "подседельный"]),
    ("Хомуты", ["хомут", "clamp"]),
    ("Вилки", ["вилка", "fork"]),
    ("Амортизаторы", ["амортиз", "shock", "suspension"]),
    ("Освещение", ["фонарь", "фара", "свет", "light", "lamp"]),
    ("Крылья", ["крыл", "fender", "mudguard"]),
    ("Багажники", ["багажник", "rack", "carrier"]),
    ("Подножки", ["поднож", "kickstand", "stand"]),
    ("Зеркала", ["зеркал", "mirror"]),
    ("Замки", ["замок", "lock"]),
    ("Крепёж", ["болт", "гайк", "винт", "крепеж", "проставка", "washer", "bolt", "nut", "screw"]),
    ("Инструменты", ["ключ", "отвёртк", "отвертк", "инструмент", "tool", "wrench"]),
    ("Смазки и химия", ["смазк", "масло", "lubric", "grease", "oil"]),
    ("Защита", ["шлем", "налокотник", "наколенник", "защит", "helmet"]),
    ("Сумки и флягодержатели", ["сумк", "рюкзак", "флягодерж", "bottle cage", "bag"]),
    ("Аксессуары", ["аксессуар", "велокомпьютер", "одометр", "звонок", "держатель телефона"]),
]

DIAMETER_RE = re.compile(r'(?<!\d)(12|14|16|18|20|24|26|27(?:[\.,]5)?|28|29)\s*(?:"|\'\'|д|inch|in)?', re.I)
VOLTAGE_RE = re.compile(r'\b(24|36|48|52|60|72)\s*v\b', re.I)
AH_RE = re.compile(r'\b(\d{1,2}(?:[\.,]\d+)?)\s*ah\b', re.I)
WATT_RE = re.compile(r'\b(\d{2,5})\s*(?:ватт?|w)\b', re.I)
MM_RE = re.compile(r'\b(120|140|160|180|203)\s*мм\b', re.I)
SPEED_RE = re.compile(r'\b([6789]|1[0-2])\s*(?:ск|скор|скорост|speed)\b', re.I)
APPAREL_SIZE_RE = re.compile(r"\b(?:xxs|xs|s|m|l|xl|xxl|xxxl)\b", re.I)
CACHE_DB = Path("photo_cache.db")

def init_state():
    defaults = {"offers_by_supplier": {}, "images_by_supplier": {}, "master_df": pd.DataFrame(), "mapping_df": pd.DataFrame(), "r2_exists_cache": {}}
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

def get_cache_conn():
    conn = sqlite3.connect(CACHE_DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_cache_db():
    conn = get_cache_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS photo_cache (
            cache_key TEXT PRIMARY KEY,
            supplier TEXT,
            supplier_article TEXT,
            normalized_name TEXT,
            photo_ref TEXT,
            r2_url TEXT,
            source_image_url TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_photo_cache_supplier_article ON photo_cache(supplier, supplier_article)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_photo_cache_normalized_name ON photo_cache(normalized_name)")
    conn.commit()
    conn.close()

def build_cache_key(supplier: str, supplier_article: str, normalized_name: str) -> str:
    supplier_article = str(supplier_article or "").strip()
    normalized_name = str(normalized_name or "").strip()
    if supplier_article:
        return f"{supplier}|article|{supplier_article}"
    return f"{supplier}|name|{normalized_name}"

def get_cached_photo(cache_key: str):
    conn = get_cache_conn()
    row = conn.execute("SELECT cache_key, supplier, supplier_article, normalized_name, photo_ref, r2_url, source_image_url FROM photo_cache WHERE cache_key = ?", (cache_key,)).fetchone()
    conn.close()
    return dict(row) if row else None

def upsert_cached_photo(cache_key: str, supplier: str, supplier_article: str, normalized_name: str, photo_ref: str, r2_url: str, source_image_url: str):
    conn = get_cache_conn()
    conn.execute("""
        INSERT INTO photo_cache(cache_key, supplier, supplier_article, normalized_name, photo_ref, r2_url, source_image_url, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(cache_key) DO UPDATE SET
            supplier = excluded.supplier,
            supplier_article = excluded.supplier_article,
            normalized_name = excluded.normalized_name,
            photo_ref = excluded.photo_ref,
            r2_url = excluded.r2_url,
            source_image_url = excluded.source_image_url,
            updated_at = CURRENT_TIMESTAMP
    """, (cache_key, supplier, supplier_article, normalized_name, photo_ref, r2_url, source_image_url))
    conn.commit()
    conn.close()

def get_photo_cache_count():
    conn = get_cache_conn()
    count = conn.execute("SELECT COUNT(*) FROM photo_cache").fetchone()[0]
    conn.close()
    return count

def normalize_name(value: str) -> str:
    value = str(value or "").lower().strip()
    value = re.sub(r"[\"'`]", "", value)
    value = re.sub(r"[^\w\s\.\-\/а-яА-Я]", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()

def signature_name(value: str) -> str:
    value = normalize_name(value)
    tokens = [t for t in value.split() if t not in {"для", "и", "в", "на", "с", "под", "комплект", "шт", "шт.", "новый", "новая"}]
    return " ".join(tokens)

def slugify_text(value: str) -> str:
    value = normalize_name(value)
    value = re.sub(r"[^\wа-яА-Я]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value[:80]

def extract_specs(text: str) -> dict:
    text = str(text or "")
    norm = normalize_name(text)
    voltage = ah = watt = diameter = rotor_mm = speeds = None
    m = VOLTAGE_RE.search(norm)
    if m:
        voltage = m.group(1)
    m = AH_RE.search(norm)
    if m:
        ah = m.group(1).replace(",", ".")
    m = WATT_RE.search(norm)
    if m:
        watt = m.group(1)
    m = DIAMETER_RE.search(norm)
    if m:
        diameter = m.group(1).replace(",", ".")
    m = MM_RE.search(norm)
    if m:
        rotor_mm = m.group(1)
    m = SPEED_RE.search(norm)
    if m:
        speeds = m.group(1)
    return {"voltage": voltage, "ah": ah, "watt": watt, "diameter": diameter, "rotor_mm": rotor_mm, "speeds": speeds}

def to_float(value):
    if pd.isna(value) or value == "":
        return None
    value = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(value)
    except Exception:
        return None

def normalize_stock(value):
    if pd.isna(value) or value == "":
        return None
    text = str(value).strip().lower()
    mapping = {"нет": 0, "мало": 1, "много": 10, "более 10": 10, "скоро будут": 0, "поз заказ": 0, "распродажа": 1}
    if text in mapping:
        return mapping[text]
    try:
        return int(float(text))
    except Exception:
        return None

def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df

def is_complete_bike(norm: str) -> bool:
    if not norm:
        return False
    bike_keywords = [
        "велосипед", "bike", "bmx", "mtb", "шоссейн", "горный велосипед",
        "детский велосипед", "складной велосипед", "электровелосипед", "e-bike", "ebike"
    ]
    exclude_keywords = [
        "рама", "колес", "покрыш", "камера", "цеп", "кассет", "переключател",
        "тормоз", "седл", "вилка", "руль", "манетк", "багажник", "крыл", "фонарь",
        "поднож", "насос", "сумк", "фляг", "флягодерж", "замок", "зеркал"
    ]
    return any(k in norm for k in bike_keywords) and not any(k in norm for k in exclude_keywords)

def infer_categories(name: str):
    raw = str(name or "")
    norm = f" {normalize_name(raw)} "
    specs = extract_specs(raw)
    diameter = specs.get("diameter")
    rotor_mm = specs.get("rotor_mm")
    speeds = specs.get("speeds")

    def has(*keywords):
        return any(k in norm for k in keywords)

    def diameter_label(prefix="Диаметр"):
        if not diameter:
            return ""
        d = str(diameter).replace(",", ".")
        if d.lower() == "700c":
            return f"{prefix} 700C"
        if d.endswith(".0"):
            d = d[:-2]
        return f'{prefix} {d}"'

    if has(" самокат ", " самоката ", " трюковый самокат", "для самоката"):
        if has("дека", "руль", "вилка", "ось", "колес", "колёса", "зажим", "хомут", "подшипник"):
            return "Запчасти для самокатов", "Комплектующие"
        return "Самокаты", "Самокаты"

    if has(" скейтборд", " пенниборд", " лонгборд", "для скейтборда", "скейтборда "):
        if has("колес", "подшипник", "ось"):
            return "Запчасти для скейтбордов", "Комплектующие"
        return "Скейтборды", "Скейтборды"

    if has(" беговел "):
        return "Беговелы", (diameter_label() or "Беговелы")

    if is_complete_bike(norm) and not has("щетка", "щётка", "бутылоч", "аптечка", "заплатк", "держатель", "замок", "крыл", "поднож", "фонарь", "флягодерж", "насос", "перчат", "сумк"):
        return "Велосипеды", (diameter_label() or "Велосипеды")

    if has(" аккумулятор", " акб ", "для акб", " battery ", " lifepo4 ", " li ion battery", " li-ion battery", "батарея") and not has("фонарь", "фара", "свет"):
        if specs["voltage"] and specs["ah"]:
            return "Аккумуляторы", f"{specs['voltage']}V {specs['ah']}Ah"
        if specs["voltage"]:
            return "Аккумуляторы", f"{specs['voltage']}V"
        return "Аккумуляторы", "АКБ"

    if has("зарядн", "зарядка", " charger ", "зарядное устройство", "блок питания"):
        return "Зарядные устройства", (f"{specs['voltage']}V" if specs["voltage"] else "Зарядки")

    if has("контроллер", " controller "):
        if specs["voltage"] and specs["watt"]:
            return "Контроллеры", f"{specs['voltage']}V {specs['watt']}W"
        if specs["voltage"]:
            return "Контроллеры", f"{specs['voltage']}V"
        return "Контроллеры", "Контроллеры"

    if has("дисплей", "панель управления", " cycle analyst ", " display ", " lcd ", " led "):
        return "Дисплеи и панели управления", "Дисплеи"

    if has("мотор-колес", "мотор колес", "мотор-втулк", "hub motor", "mid drive", "редукторный мотор", "электромотор"):
        if specs["voltage"] and specs["watt"]:
            return "Моторы", f"{specs['voltage']}V {specs['watt']}W"
        if specs["watt"]:
            return "Моторы", f"{specs['watt']}W"
        return "Моторы", "Моторы"

    if has("кабель", "провод", "разъём", "разъем", "коннектор", "connector", "wire", "cable", "удлинитель", "датчик скорости", "сенсор тормоза", "штекер"):
        return "Кабели и разъёмы", "Кабели и разъёмы"

    if has("флягодерж", "bottle cage", "держатель для бутыл"):
        return "Флягодержатели", "Флягодержатели"

    if has("бутылоч", "бутылка", "фляга", "термофляга"):
        return "Фляги", "Фляги"

    if has("сумк", "рюкзак", "bikepacking", "велосумка", "чехол на раму"):
        if has("под седло", "подседел"):
            return "Сумки", "Подседельные"
        if has("на раму", "рамная"):
            return "Сумки", "На раму"
        if has("на руль", "нарульная"):
            return "Сумки", "На руль"
        return "Сумки", "Сумки"

    if has("держатель для смартфона", "держатель смартфона", "держатель телефона", "phone holder"):
        return "Держатели для телефона", "Держатели"

    if has("звонок", "клаксон", "сигнал", "сирена"):
        return "Звонки и сигналы", "Звонки и сигналы"

    if has("покрыш", "шина", "tire", "tyre") and not has("ободная лента"):
        return "Покрышки", (diameter_label() or "Покрышки")

    if has("камера", "tube", "innertube", "флиппер"):
        return "Камеры", (diameter_label() or "Камеры")

    if has("ободная лента", "обод", " rim "):
        return "Обода и спицы", (diameter_label("Обод") or "Обода")

    if has("спиц", "ниппель спицы", " spoke "):
        return "Обода и спицы", "Спицы"

    if has("втулк", " hub "):
        if has("перед"):
            return "Втулки", "Передние"
        if has("зад"):
            return "Втулки", "Задние"
        return "Втулки", "Втулки"

    if has("эксцентрик", "ось колеса", "ось переднего", "ось заднего"):
        return "Оси и эксцентрики", "Оси и эксцентрики"

    if has("колесо в сборе", "wheelset", "вилсет") or (" колесо " in norm and not has("держатель велосипеда", "крепление за колесо", "колесико переключателя", "ролик переключателя", "для самоката")):
        if has("перед"):
            return "Колёса", (diameter_label("Переднее") or "Передние")
        if has("зад"):
            return "Колёса", (diameter_label("Заднее") or "Задние")
        return "Колёса", (diameter_label() or "Колёса")

    if has("колодк", "brake pad"):
        return "Тормозные колодки", "Колодки"

    if has("ротор", "диск тормоз", "rotor", "disc brake rotor"):
        return "Роторы", (f"{rotor_mm} мм" if rotor_mm else "Роторы")

    if has("тормоз", "brake", "гидролиния", "калипер"):
        if "гидр" in norm:
            return "Тормоза", "Гидравлические"
        if "мех" in norm:
            return "Тормоза", "Механические"
        if has("v-brake", "ободной", "cantilever"):
            return "Тормоза", "Ободные"
        return "Тормоза", "Тормоза"

    if has("цепь"):
        return "Цепи", (f"{speeds}-скоростная" if speeds else "Цепи")

    if has("кассет", "cassette"):
        return "Кассеты и трещотки", (f"Кассеты {speeds}-скоростные" if speeds else "Кассеты")

    if has("трещот", "freewheel"):
        return "Кассеты и трещотки", (f"Трещотки {speeds}-скоростные" if speeds else "Трещотки")

    if has("звезд", "звезда", "sprocket", "chainring"):
        return "Звёзды", "Звёзды"

    if has("шатун", "crank"):
        return "Шатуны и каретки", "Шатуны"

    if has("каретк", "bottom bracket"):
        return "Шатуны и каретки", "Каретки"

    if has("педал", "pedal"):
        return "Педали", "Педали"

    if has("переключател", "derailleur", "ролик переключателя", "колесико переключателя"):
        if has("ролик", "колесико"):
            return "Переключатели", "Ролики"
        if has("перед"):
            return "Переключатели", "Передние"
        if has("зад"):
            return "Переключатели", "Задние"
        return "Переключатели", "Переключатели"

    if has("манетк", "gripshift", "trigger shifter", "шифтер"):
        return "Манетки", "Манетки"

    if has("рулевая колонка", "рулевая", "headset"):
        return "Рулевые колонки", "Рулевые колонки"

    if has("руль", "handlebar", "guidon") and not has("держатель"):
        return "Рули", "Рули"

    if has("вынос", "stem"):
        return "Выносы", "Выносы"

    if has("грипс", "грипсы", "ручк руля", " grip ", "обмотка руля", "bar tape"):
        if has("обмотка"):
            return "Грипсы", "Обмотка руля"
        return "Грипсы", "Грипсы"

    if has("седл", "saddle") and not has("под седло", "подседельная"):
        return "Седла", "Седла"

    if has("подседел", "seatpost"):
        return "Подседельные штыри", "Подседельные штыри"

    if has("хомут", "clamp"):
        return "Хомуты", "Хомуты"

    if has("вилка", "fork") and not has("якорь вилки", "держатель"):
        return "Вилки", (diameter_label() or "Вилки")

    if has("амортиз", "shock", "suspension"):
        return "Амортизаторы", "Амортизаторы"

    if has("насос", "pump", "головка насоса", "шланг для насоса", "зажим для шланга"):
        if has("напольн", "ножн"):
            return "Насосы", "Напольные"
        if has("ручн", "мини"):
            return "Насосы", "Ручные"
        if has("головка", "шланг", "зажим"):
            return "Насосы", "Запчасти и аксессуары"
        return "Насосы", "Насосы"

    if has("крыл", "брызговик", "fender", "mudguard"):
        return "Крылья", (diameter_label() or "Крылья")

    if has("багажник", "rack", "carrier"):
        return "Багажники", "Багажники"

    if has("поднож", "kickstand", "stand"):
        return "Подножки", "Подножки"

    if has("зеркал", "mirror"):
        return "Зеркала", "Зеркала"

    if has("замок", " lock ", "противоугон"):
        return "Замки", "Замки"

    if has("фонарь", "фара", "light", "lamp", "свет") and not has("светоотраж", "светоотражатель"):
        if has("зад"):
            return "Освещение", "Задние"
        if has("перед"):
            return "Освещение", "Передние"
        if has("комплект"):
            return "Освещение", "Комплекты"
        return "Освещение", "Освещение"

    if has("светоотраж", "катафот", "рефлектор"):
        return "Светоотражатели", "Светоотражатели"

    if has("ключ", "отвёртк", "отвертк", "инструмент", "tool", "wrench", "шестигран", "мультитул", "съемник", "съёмник", "монтажк", "выжимка цепи", "ремкомплект", "аптечка", "заплатк"):
        if has("монтажк"):
            return "Инструменты", "Монтажки"
        if has("выжимка цепи"):
            return "Инструменты", "Выжимки цепи"
        if has("ремкомплект", "аптечка", "заплатк"):
            return "Инструменты", "Ремкомплекты"
        return "Инструменты", "Инструменты"

    if has("смазк", "масло", "lubric", "grease", "oil", "герметик", "очиститель"):
        return "Смазки и химия", "Смазки и химия"

    size_match = APPAREL_SIZE_RE.search(norm)
    if has("велоперчат", "перчат", "шлем", "налокотник", "наколенник", "защит", "helmet", "бахил", "glove"):
        if has("перчат", "glove"):
            return "Экипировка и защита", (f"Перчатки {size_match.group(0).upper()}" if size_match else "Перчатки")
        if has("бахил"):
            return "Экипировка и защита", "Бахилы"
        if has("шлем", "helmet"):
            return "Экипировка и защита", "Шлемы"
        return "Экипировка и защита", "Защита"

    if has("болт", "гайк", "винт", "крепеж", "крепёж", "проставка", "washer", "bolt", "nut", "screw", "шайб"):
        return "Крепёж", "Крепёж"

    return "Прочее", ""


def normalize_google_sheet_url(url: str) -> str:
    if "docs.google.com/spreadsheets" in url and "/edit" in url:
        return url.split("/edit")[0] + "/export?format=xlsx"
    return url

def read_source_bytes(source_type, uploaded_file, source_url):
    if source_type == "file":
        if uploaded_file is None:
            raise ValueError("Нужно загрузить файл.")
        return uploaded_file.name, uploaded_file.getvalue()
    if not source_url.strip():
        raise ValueError("Нужно указать ссылку.")
    final_url = normalize_google_sheet_url(source_url.strip())
    response = requests.get(final_url, timeout=60)
    response.raise_for_status()
    filename = final_url.split("/")[-1] or "supplier.xlsx"
    if "format=xlsx" in final_url and not filename.endswith(".xlsx"):
        filename = "supplier.xlsx"
    return filename, response.content

def load_source_to_df(filename, file_bytes, sheet_name, header_row):
    suffix = Path(filename).suffix.lower()
    workbook = None
    if suffix == ".csv":
        df = pd.read_csv(io.BytesIO(file_bytes))
    elif suffix == ".xls":
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_row, engine="xlrd")
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_row)
        workbook = load_workbook(io.BytesIO(file_bytes))
    df = clean_columns(df)
    df["__excel_row__"] = range(header_row + 2, header_row + 2 + len(df))
    return df, workbook

def extract_hyperlinks_map(ws, header_row):
    mapping = {}
    excel_header_row = header_row + 1
    photo_col = None
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(excel_header_row, col).value).strip() == "Товар на сайте":
            photo_col = col
            break
    if photo_col is None:
        return mapping
    for row in range(excel_header_row + 1, ws.max_row + 1):
        cell = ws.cell(row, photo_col)
        if cell.hyperlink and cell.hyperlink.target:
            mapping[row] = cell.hyperlink.target
    return mapping

def build_stable_photo_seed(supplier_key: str, supplier_article: str, name: str, normalized_name: str = "") -> str:
    article = str(supplier_article or "").strip()
    normalized_name = normalized_name or normalize_name(name)
    if article:
        return f"{supplier_key}|article|{article.lower()}"
    signature = signature_name(normalized_name) or slugify_text(normalized_name) or "item"
    return f"{supplier_key}|name|{signature}"

def safe_ext_from_url(url: str) -> str:
    for ext in [".jpg", ".jpeg", ".png", ".webp"]:
        if ext in url.lower():
            return ext
    return ".jpg"

def extract_images_map(ws):
    images = {}
    for img in getattr(ws, "_images", []):
        try:
            excel_row = img.anchor._from.row + 1
            data = img._data()
            images[excel_row] = data
        except Exception:
            continue
    return images

def download_image_bytes(url: str):
    try:
        resp = requests.get(url, timeout=20)
        resp.raise_for_status()
        return resp.content
    except Exception:
        return None

def build_photo_ref(prefix: str, seed: str, ext: str):
    digest = hashlib.md5(seed.encode("utf-8")).hexdigest()[:16]
    return f"images/{prefix}_{digest}{ext}"

def attach_images(parsed_df, supplier_key, workbook, header_row):
    parsed_df = parsed_df.copy()
    image_store = {}
    photo_refs, source_urls = [], []
    if supplier_key == "s2" and workbook is not None:
        ws = workbook[SUPPLIERS[supplier_key]["sheet_name"]]
        links_map = extract_hyperlinks_map(ws, header_row)
        for _, row in parsed_df.iterrows():
            excel_row = int(row.get("__excel_row__", 0))
            url = links_map.get(excel_row, "")
            source_urls.append(url)
            if url:
                ext = safe_ext_from_url(url)
                seed = build_stable_photo_seed(supplier_key, row.get("supplier_article", ""), row.get("name", ""), row.get("normalized_name", ""))
                photo_ref = build_photo_ref(supplier_key, seed, ext)
                photo_refs.append(photo_ref)
            else:
                photo_refs.append("")
        parsed_df["source_image_url"] = source_urls
        parsed_df["photo_ref"] = photo_refs
        return parsed_df, {}
    if supplier_key in ["s3", "s4"] and workbook is not None:
        ws = workbook[SUPPLIERS[supplier_key]["sheet_name"]]
        img_map = extract_images_map(ws)
        for _, row in parsed_df.iterrows():
            excel_row = int(row.get("__excel_row__", 0))
            img_bytes = img_map.get(excel_row)
            if img_bytes:
                seed = build_stable_photo_seed(supplier_key, row.get("supplier_article", ""), row.get("name", ""), row.get("normalized_name", ""))
                photo_ref = build_photo_ref(supplier_key, seed, ".png")
                image_store[photo_ref] = img_bytes
                photo_refs.append(photo_ref)
            else:
                photo_refs.append("")
            source_urls.append("")
        parsed_df["source_image_url"] = source_urls
        parsed_df["photo_ref"] = photo_refs
        return parsed_df, image_store
    parsed_df["source_image_url"] = parsed_df.get("image_url", "")
    parsed_df["photo_ref"] = ""
    return parsed_df, image_store

def apply_selected_price_tier(df, selected_tier: str):
    df = df.copy()
    if selected_tier in df.columns:
        df["base_price"] = df[selected_tier]
        df["price_tier"] = selected_tier
    else:
        df["base_price"] = df.get("price")
        df["price_tier"] = "price"
    return df

def enrich_categories(df):
    df = df.copy()
    cats = df["name"].apply(infer_categories)
    df["category_l1"] = [x[0] for x in cats]
    df["category_l2"] = [x[1] for x in cats]
    return df

def parse_supplier1(df):
    df = df.rename(columns={"Артикул": "supplier_article", "Наименование товара": "name", "Ед.": "unit", "в уп": "pack_qty", "цена": "price", "фото": "photo"}).copy()
    if "supplier_article" not in df.columns or "name" not in df.columns:
        raise ValueError("Не найдены нужные колонки для поставщика 1")
    df["supplier_article"] = df["supplier_article"].fillna("").astype(str).str.strip()
    df["name"] = df["name"].fillna("").astype(str).str.strip()
    df = df[(df["supplier_article"] != "") & (df["name"] != "")]
    df = df[~df["name"].str.startswith("(")]
    df["price"] = df["price"].apply(to_float)
    df["stock"] = None
    df["image_url"] = ""
    df["supplier"] = "s1"
    return df[["supplier", "supplier_article", "name", "price", "stock", "image_url", "__excel_row__"]]

def parse_supplier2(df):
    df = df.rename(columns={"Артикул": "supplier_article", "Номенклатура": "name", "Остаток, шт.": "stock", "Цена Опт 1, руб.": "price", "Цена Опт 2, руб.": "price_opt2", "Цена Опт 3, руб.": "price_opt3", "Цена РРЦ, руб.": "price_rrc"}).copy()
    if "supplier_article" not in df.columns or "name" not in df.columns:
        raise ValueError("Не найдены нужные колонки для поставщика 2")
    df["supplier_article"] = df["supplier_article"].fillna("").astype(str).str.strip()
    df["name"] = df["name"].fillna("").astype(str).str.strip()
    df = df[(df["supplier_article"] != "") & (df["name"] != "")]
    df["stock"] = df["stock"].apply(normalize_stock)
    for col in ["price", "price_opt2", "price_opt3", "price_rrc"]:
        if col in df.columns:
            df[col] = df[col].apply(to_float)
    df["image_url"] = ""
    df["supplier"] = "s2"
    return df[["supplier", "supplier_article", "name", "price", "price_opt2", "price_opt3", "price_rrc", "stock", "image_url", "__excel_row__"]]

def parse_supplier3(df):
    df = df.rename(columns={"Наименование товара": "name", "Фото": "photo", "Цена, ₽": "price", "От 10шт": "price_opt10", "От 50шт": "price_opt50", "Безнал": "payment_note"}).copy()
    if "name" not in df.columns:
        raise ValueError("Не найдена колонка 'Наименование товара' для поставщика 3")
    df["name"] = df["name"].fillna("").astype(str).str.strip()
    df = df[df["name"] != ""]
    df = df[~df["name"].str.contains("WhatsApp|Наличие и актуальные цены", case=False, na=False)]
    for col in ["price", "price_opt10", "price_opt50"]:
        if col in df.columns:
            df[col] = df[col].apply(to_float)
    df["supplier_article"] = ""
    df["stock"] = None
    df["image_url"] = ""
    df["supplier"] = "s3"
    return df[["supplier", "supplier_article", "name", "price", "price_opt10", "price_opt50", "stock", "image_url", "__excel_row__"]]

def parse_supplier4(df):
    first_col = df.columns[0]
    df = df.rename(columns={first_col: "row_num", "Название": "name", "Наличие": "stock", "упаковка (шт)": "pack_qty", "вес 1 шт (кг)": "weight_kg", "РРЦ": "price_rrc", "от 1 уп": "price", "Своим": "own_price", "Изображение": "photo"}).copy()
    if "name" not in df.columns:
        raise ValueError("Не найдена колонка 'Название' для поставщика 4")
    df["name"] = df["name"].fillna("").astype(str).str.strip()
    df = df[df["name"] != ""]
    df["stock"] = df["stock"].apply(normalize_stock)
    for col in ["price", "price_rrc", "own_price", "pack_qty", "weight_kg"]:
        if col in df.columns:
            df[col] = df[col].apply(to_float)
    df["supplier_article"] = ""
    df["image_url"] = ""
    df["supplier"] = "s4"
    return df[["supplier", "supplier_article", "name", "price", "price_rrc", "own_price", "pack_qty", "weight_kg", "stock", "image_url", "__excel_row__"]]

def parse_supplier(supplier_key, df):
    if supplier_key == "s1": return parse_supplier1(df)
    if supplier_key == "s2": return parse_supplier2(df)
    if supplier_key == "s3": return parse_supplier3(df)
    if supplier_key == "s4": return parse_supplier4(df)
    raise ValueError("Неизвестный поставщик")

def add_normalized_columns(df):
    df = df.copy()
    df["normalized_name"] = df["name"].apply(normalize_name)
    df["name_signature"] = df["name"].apply(signature_name)
    df["specs"] = df["name"].apply(extract_specs)
    return df

def duplicate_score(row_a: dict, row_b: dict) -> float:
    a = row_a["name_signature"]
    b = row_b["name_signature"]
    score = max(fuzz.token_sort_ratio(a, b), fuzz.token_set_ratio(a, b), fuzz.partial_ratio(a, b))
    specs_a = row_a.get("specs", {}) or {}
    specs_b = row_b.get("specs", {}) or {}
    for key in ["voltage", "ah", "watt", "diameter", "rotor_mm", "speeds"]:
        va = specs_a.get(key); vb = specs_b.get(key)
        if va and vb:
            if va == vb: score += 4
            else: score -= 10
    if row_a.get("category_l1") == row_b.get("category_l1"): score += 2
    if row_a.get("category_l2") and row_a.get("category_l2") == row_b.get("category_l2"): score += 2
    return min(score, 100)

def build_master(offers_df):
    if offers_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    offers_df = offers_df.copy()
    offers_df["normalized_name"] = offers_df["name"].apply(normalize_name)
    offers_df["name_signature"] = offers_df["name"].apply(signature_name)
    offers_df["specs"] = offers_df["name"].apply(extract_specs)
    offers_df = enrich_categories(offers_df)
    master_rows, mapping_rows = [], []
    master_id = 1
    masters = []

    def create_master_from_group(grp: pd.DataFrame, method: str, confidence: float):
        nonlocal master_id, master_rows, mapping_rows, masters
        best_price = grp["base_price"].dropna().min() if "base_price" in grp.columns else None
        stock_sum = grp["stock"].dropna().sum() if "stock" in grp.columns else None
        photo_ref = next((x for x in grp["photo_ref"].fillna("").tolist() if x), "")
        representative_idx = grp["base_price"].fillna(float("inf")).idxmin() if grp["base_price"].notna().any() else grp.index[0]
        representative = grp.loc[representative_idx]
        row = {
            "master_id": master_id,
            "article": str(representative.get("supplier_article", "") or ""),
            "name": representative["name"],
            "normalized_name": representative["normalized_name"],
            "category_l1": representative.get("category_l1", "Прочее"),
            "category_l2": representative.get("category_l2", ""),
            "final_price": best_price,
            "final_stock": int(stock_sum) if pd.notna(stock_sum) else None,
            "final_image": photo_ref,
        }
        master_rows.append(row)
        masters.append({
            "master_id": master_id,
            "normalized_name": representative["normalized_name"],
            "name_signature": representative["name_signature"],
            "specs": representative["specs"],
            "category_l1": representative.get("category_l1", ""),
            "category_l2": representative.get("category_l2", ""),
        })
        for _, item in grp.iterrows():
            mapping_rows.append({
                "supplier": item["supplier"], "supplier_article": item.get("supplier_article", ""), "supplier_name": item["name"],
                "normalized_name": item["normalized_name"], "category_l1": item.get("category_l1", ""), "category_l2": item.get("category_l2", ""),
                "master_id": master_id, "match_method": method, "confidence": confidence,
            })
        master_id += 1

    article_df = offers_df[offers_df["supplier_article"].fillna("") != ""].copy()
    no_article_df = offers_df[offers_df["supplier_article"].fillna("") == ""].copy()
    if not article_df.empty:
        for article, grp in article_df.groupby("supplier_article"):
            create_master_from_group(grp, "article_exact", 100.0)
    exact_name_groups = {}
    if not no_article_df.empty:
        for idx, row in no_article_df.iterrows():
            sig = row["name_signature"]
            exact_name_groups.setdefault(sig, []).append(idx)
        used = set()
        for sig, idxs in exact_name_groups.items():
            if len(idxs) >= 2:
                grp = no_article_df.loc[idxs]
                create_master_from_group(grp, "name_exact_signature", 99.0)
                used.update(idxs)
        no_article_df = no_article_df.drop(index=list(used), errors="ignore")
    for _, row in no_article_df.iterrows():
        best_master_id = None
        best_score = -1
        row_payload = {"name_signature": row["name_signature"], "specs": row["specs"], "category_l1": row.get("category_l1", ""), "category_l2": row.get("category_l2", "")}
        for master in masters:
            score = duplicate_score(row_payload, master)
            if score > best_score:
                best_score = score
                best_master_id = master["master_id"]
        if best_score >= 90 and best_master_id is not None:
            mapping_rows.append({
                "supplier": row["supplier"], "supplier_article": row.get("supplier_article", ""), "supplier_name": row["name"],
                "normalized_name": row["normalized_name"], "category_l1": row.get("category_l1", ""), "category_l2": row.get("category_l2", ""),
                "master_id": best_master_id, "match_method": "name_fuzzy_strong", "confidence": float(best_score),
            })
            for mr in master_rows:
                if mr["master_id"] == best_master_id:
                    if pd.notna(row.get("base_price")):
                        if pd.isna(mr["final_price"]) or float(row["base_price"]) < float(mr["final_price"]):
                            mr["final_price"] = float(row["base_price"])
                            mr["name"] = row["name"]
                            mr["normalized_name"] = row["normalized_name"]
                            mr["category_l1"] = row.get("category_l1", mr["category_l1"])
                            mr["category_l2"] = row.get("category_l2", mr["category_l2"])
                    if pd.notna(row.get("stock")):
                        existing = mr.get("final_stock")
                        mr["final_stock"] = (existing or 0) + int(row["stock"])
                    if not mr.get("final_image") and row.get("photo_ref"):
                        mr["final_image"] = row.get("photo_ref")
                    break
        else:
            create_master_from_group(pd.DataFrame([row]), "new_name", 100.0)
    master_df = pd.DataFrame(master_rows)
    mapping_df = pd.DataFrame(mapping_rows)
    if not master_df.empty:
        master_df = master_df.sort_values(["category_l1", "category_l2", "name"]).reset_index(drop=True)
    if not mapping_df.empty:
        mapping_df = mapping_df.sort_values(["match_method", "confidence"], ascending=[True, False]).reset_index(drop=True)
    return master_df, mapping_df

def build_excel_bytes(df: pd.DataFrame, sheet_name: str):
    output = io.BytesIO()
    export_df = df.copy()
    if sheet_name == "final_price":
        rename_map = {
            "master_id": "ID",
            "article": "Артикул",
            "normalized_name": "Наименование",
            "category_l1": "Категория",
            "category_l2": "Подкатегория",
            "final_price": "Закупка",
            "final_stock": "Наличие",
            "price_with_markup": "Цена Оптовая",
            "final_image_public_url": "Ссылка на фото",
        }
        ordered_cols = [
            "master_id", "article", "normalized_name", "category_l1", "category_l2",
            "final_price", "final_stock", "price_with_markup", "final_image_public_url",
        ]
        for col in ordered_cols:
            if col not in export_df.columns:
                export_df[col] = ""
        export_df = export_df[ordered_cols].rename(columns=rename_map)
    for col in export_df.columns:
        if export_df[col].dtype == "object":
            export_df[col] = export_df[col].fillna("")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output.getvalue()

def has_r2_config():
    required = ["R2_ACCOUNT_ID", "R2_ACCESS_KEY_ID", "R2_SECRET_ACCESS_KEY", "R2_BUCKET_NAME", "R2_PUBLIC_BASE_URL"]
    return all(key in st.secrets for key in required)

def get_r2_client():
    return boto3.client(
        service_name="s3",
        endpoint_url=f"https://{st.secrets['R2_ACCOUNT_ID']}.r2.cloudflarestorage.com",
        aws_access_key_id=st.secrets["R2_ACCESS_KEY_ID"],
        aws_secret_access_key=st.secrets["R2_SECRET_ACCESS_KEY"],
        region_name="auto",
    )

def r2_public_url_for_key(key: str) -> str:
    return f"{st.secrets['R2_PUBLIC_BASE_URL'].rstrip('/')}/{key}"

def r2_object_exists(key: str) -> bool:
    cache = st.session_state.r2_exists_cache
    if key in cache:
        return cache[key]
    s3 = get_r2_client()
    try:
        s3.head_object(Bucket=st.secrets["R2_BUCKET_NAME"], Key=key)
        cache[key] = True
        return True
    except ClientError as e:
        code = str(e.response.get("Error", {}).get("Code", ""))
        if code in ["404", "NoSuchKey", "NotFound"]:
            cache[key] = False
            return False
        raise

def upload_bytes_to_r2_if_needed(key: str, data: bytes, content_type: str = "application/octet-stream"):
    if r2_object_exists(key):
        return r2_public_url_for_key(key), "cached"
    s3 = get_r2_client()
    s3.put_object(Bucket=st.secrets["R2_BUCKET_NAME"], Key=key, Body=data, ContentType=content_type)
    st.session_state.r2_exists_cache[key] = True
    return r2_public_url_for_key(key), "uploaded"

def guess_content_type(key: str) -> str:
    key = key.lower()
    if key.endswith(".png"): return "image/png"
    if key.endswith(".webp"): return "image/webp"
    if key.endswith(".jpeg") or key.endswith(".jpg"): return "image/jpeg"
    return "application/octet-stream"

def upload_final_images_to_r2(export_df: pd.DataFrame, offers_by_supplier: dict, images_by_supplier: dict):
    export_df = export_df.copy()
    uploaded_map = {}
    stats = {"cache_db": 0, "cache_r2": 0, "uploaded": 0, "missed": 0}
    if not has_r2_config():
        raise ValueError("R2 secrets не настроены в Streamlit.")
    in_memory = {}
    source_lookup = {}
    for supplier_images in images_by_supplier.values():
        in_memory.update(supplier_images)
    for _, df in offers_by_supplier.items():
        if df is None or df.empty:
            continue
        for _, row in df.iterrows():
            supplier = str(row.get("supplier", "") or "")
            supplier_article = str(row.get("supplier_article", "") or "")
            normalized_name = str(row.get("normalized_name", "") or "")
            cache_key = build_cache_key(supplier, supplier_article, normalized_name)
            source_lookup[cache_key] = {
                "source_image_url": str(row.get("source_image_url", "") or ""),
                "photo_ref": str(row.get("photo_ref", "") or ""),
                "supplier": supplier,
                "supplier_article": supplier_article,
                "normalized_name": normalized_name,
            }
    final_urls = []
    progress = st.progress(0, text="Проверка photo-cache и загрузка фото в R2...")
    total = max(len(export_df), 1)
    for idx, (_, row) in enumerate(export_df.iterrows(), start=1):
        key = str(row.get("final_image", "") or "").strip()
        article = str(row.get("article", "") or "")
        master_norm = str(row.get("normalized_name", "") or "")
        if not key:
            final_urls.append("")
            stats["missed"] += 1
            progress.progress(idx / total, text=f"Фото {idx}/{total}")
            continue
        if key in uploaded_map:
            final_urls.append(uploaded_map[key])
            progress.progress(idx / total, text=f"Фото {idx}/{total}")
            continue
        found_url = None
        found_cache_key = None
        for supplier_code in ["s1", "s2", "s3", "s4"]:
            ck = build_cache_key(supplier_code, article, master_norm)
            cached = get_cached_photo(ck)
            if cached and cached.get("photo_ref") == key and cached.get("r2_url"):
                found_url = cached["r2_url"]
                found_cache_key = ck
                break
        if found_url:
            uploaded_map[key] = found_url
            final_urls.append(found_url)
            stats["cache_db"] += 1
            progress.progress(idx / total, text=f"Фото {idx}/{total}")
            continue
        if r2_object_exists(key):
            public_url = r2_public_url_for_key(key)
            uploaded_map[key] = public_url
            final_urls.append(public_url)
            stats["cache_r2"] += 1
            for supplier_code in ["s1", "s2", "s3", "s4"]:
                ck = build_cache_key(supplier_code, article, master_norm)
                src = source_lookup.get(ck)
                if src and src.get("photo_ref") == key:
                    upsert_cached_photo(ck, src["supplier"], src["supplier_article"], src["normalized_name"], key, public_url, src["source_image_url"])
                    break
            progress.progress(idx / total, text=f"Фото {idx}/{total}")
            continue
        source_data = None
        for ck, payload in source_lookup.items():
            if payload["photo_ref"] == key:
                source_data = payload
                found_cache_key = ck
                break
        data = in_memory.get(key)
        if data:
            public_url, mode = upload_bytes_to_r2_if_needed(key, data, guess_content_type(key))
            uploaded_map[key] = public_url
            final_urls.append(public_url)
            stats["uploaded" if mode == "uploaded" else "cache_r2"] += 1
            if source_data and found_cache_key:
                upsert_cached_photo(found_cache_key, source_data["supplier"], source_data["supplier_article"], source_data["normalized_name"], key, public_url, source_data["source_image_url"])
            progress.progress(idx / total, text=f"Фото {idx}/{total}")
            continue
        if source_data and source_data["source_image_url"]:
            data = download_image_bytes(source_data["source_image_url"])
            if data:
                public_url, mode = upload_bytes_to_r2_if_needed(key, data, guess_content_type(key))
                uploaded_map[key] = public_url
                final_urls.append(public_url)
                stats["uploaded" if mode == "uploaded" else "cache_r2"] += 1
                if found_cache_key:
                    upsert_cached_photo(found_cache_key, source_data["supplier"], source_data["supplier_article"], source_data["normalized_name"], key, public_url, source_data["source_image_url"])
                progress.progress(idx / total, text=f"Фото {idx}/{total}")
                continue
        final_urls.append("")
        stats["missed"] += 1
        progress.progress(idx / total, text=f"Фото {idx}/{total}")
    progress.empty()
    export_df["final_image_public_url"] = final_urls
    return export_df, stats

init_state()
init_cache_db()

with st.sidebar:
    st.title("Price Aggregator")
    page = st.radio("Раздел", ["Дашборд", "Загрузка прайсов", "Дубли и склейка", "Итоговый прайс", "R2", "Photo Cache"])

if page == "Дашборд":
    st.title("🏠 Дашборд")
    supplier_stats = []
    total_rows = 0
    for key, cfg in SUPPLIERS.items():
        df = st.session_state.offers_by_supplier.get(key, pd.DataFrame())
        count = len(df) if not df.empty else 0
        total_rows += count
        supplier_stats.append({"Поставщик": cfg["label"], "Код": key, "Строк загружено": count, "Тип источника": cfg["source_type"]})
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Всего строк поставщиков", total_rows)
    c2.metric("Карточек в итоге", len(st.session_state.master_df))
    c3.metric("Связей", len(st.session_state.mapping_df))
    c4.metric("Photo-cache SQLite", get_photo_cache_count())
    st.dataframe(pd.DataFrame(supplier_stats), use_container_width=True)
    st.info("SQLite здесь только как ускоритель. Главный кэш фото — в R2: при стабильных photo_ref повторная выгрузка должна переиспользовать уже загруженные изображения.")

elif page == "Загрузка прайсов":
    st.title("📥 Загрузка прайсов")
    supplier = st.selectbox("Выберите поставщика", list(SUPPLIERS.keys()), format_func=lambda x: SUPPLIERS[x]["label"])
    cfg = SUPPLIERS[supplier]
    col1, col2 = st.columns(2)
    with col1:
        source_type = st.radio("Источник", ["file", "url"], horizontal=True)
    with col2:
        allowed_price_tiers = cfg.get("allowed_price_tiers", ["price"])
        default_price_tier = cfg.get("default_price_tier", "price")
        default_index = allowed_price_tiers.index(default_price_tier) if default_price_tier in allowed_price_tiers else 0
        selected_price_tier = st.selectbox("Основная цена", allowed_price_tiers, index=default_index, format_func=lambda x: PRICE_TIER_LABELS[x])
    uploaded_file = None
    source_url = ""
    if source_type == "file":
        uploaded_file = st.file_uploader("Загрузите файл", type=["xls", "xlsx", "csv"])
    else:
        source_url = st.text_input("Ссылка на прайс / Google Sheets")
    if st.button("Обработать прайс"):
        try:
            filename, file_bytes = read_source_bytes(source_type, uploaded_file, source_url)
            raw_df, workbook = load_source_to_df(filename, file_bytes, cfg["sheet_name"], cfg["header_row"])
            parsed = parse_supplier(supplier, raw_df)
            parsed = add_normalized_columns(parsed)
            parsed = enrich_categories(parsed)
            parsed = apply_selected_price_tier(parsed, selected_price_tier)
            parsed, image_store = attach_images(parsed, supplier, workbook, cfg["header_row"])
            st.session_state.offers_by_supplier[supplier] = parsed
            st.session_state.images_by_supplier[supplier] = image_store
            st.success("Прайс обработан.")
            st.dataframe(parsed.head(100), use_container_width=True)
            normalized_preview = parsed.drop(columns=["__excel_row__"], errors="ignore")
            excel_bytes = build_excel_bytes(normalized_preview, "normalized")
            st.download_button("Скачать нормализованный Excel", data=excel_bytes, file_name=f"{supplier}_normalized.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if supplier == "s2":
                st.info("Для Форвард СПб фото не скачиваются во время импорта. Они загрузятся в R2 только для финальных товаров и затем сохранятся в локальный photo-cache.")
        except Exception as e:
            st.error(f"Ошибка: {e}")

elif page == "Дубли и склейка":
    st.title("🔎 Дубли и склейка")
    dfs = [df for df in st.session_state.offers_by_supplier.values() if not df.empty]
    if not dfs:
        st.info("Сначала загрузите хотя бы один прайс.")
    else:
        all_offers = pd.concat(dfs, ignore_index=True)
        if st.button("Запустить склейку"):
            master_df, mapping_df = build_master(all_offers)
            st.session_state.master_df = master_df
            st.session_state.mapping_df = mapping_df
            st.success("Склейка выполнена.")
        if not st.session_state.master_df.empty:
            st.subheader("Итоговые карточки по категориям")
            st.dataframe(st.session_state.master_df, use_container_width=True)

elif page == "Итоговый прайс":
    st.title("📤 Итоговый прайс")
    master_df = st.session_state.master_df.copy()
    if master_df.empty:
        st.info("Сначала загрузите прайсы и выполните склейку.")
    else:
        markup_percent = st.number_input("Наценка, %", min_value=0.0, max_value=1000.0, value=30.0, step=1.0)
        category_l1_options = sorted([x for x in master_df["category_l1"].dropna().unique().tolist() if str(x).strip()])
        selected_l1 = st.multiselect("Какие группы выгружать", category_l1_options, default=category_l1_options)
        filtered_df = master_df.copy()
        if selected_l1:
            filtered_df = filtered_df[filtered_df["category_l1"].isin(selected_l1)]
        category_l2_options = sorted([x for x in filtered_df["category_l2"].dropna().unique().tolist() if str(x).strip()])
        selected_l2 = st.multiselect("Подгруппы (необязательно)", category_l2_options, default=[])
        if selected_l2:
            filtered_df = filtered_df[filtered_df["category_l2"].isin(selected_l2)]
        export_df = filtered_df.copy()
        export_df["price_with_markup"] = export_df["final_price"].apply(lambda x: round(float(x) * (1 + markup_percent / 100), 2) if pd.notna(x) else None)
        auto_upload = st.checkbox("Автоматически загрузить фото в R2 перед выгрузкой", value=False)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Карточек", len(export_df))
        c2.metric("С фото", int((export_df["final_image"].fillna("") != "").sum()))
        c3.metric("Наценка", f"{markup_percent:.0f}%")
        c4.metric("Photo-cache SQLite", get_photo_cache_count())
        if auto_upload:
            if has_r2_config():
                try:
                    export_df, stats = upload_final_images_to_r2(export_df, st.session_state.offers_by_supplier, st.session_state.images_by_supplier)
                    st.success(f"Photo-cache: из базы {stats['cache_db']}, уже были в R2 {stats['cache_r2']}, загружены новые {stats['uploaded']}, без фото {stats['missed']}.")
                except Exception as e:
                    st.error(f"Ошибка загрузки фото в R2: {e}")
            else:
                st.warning("R2 secrets не настроены. Открой вкладку R2.")
        st.dataframe(export_df, use_container_width=True)
        final_excel = build_excel_bytes(export_df, "final_price")
        st.download_button("Скачать итоговый Excel", data=final_excel, file_name="final_price_list.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

elif page == "R2":
    st.title("☁️ R2")
    st.write("Система сначала смотрит photo-cache, потом R2, и только потом качает фото у поставщика.")
    st.code("""R2_ACCOUNT_ID = "ваш_account_id"
R2_ACCESS_KEY_ID = "ваш_access_key_id"
R2_SECRET_ACCESS_KEY = "ваш_secret_access_key"
R2_BUCKET_NAME = "images"
R2_PUBLIC_BASE_URL = "https://pub-xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx.r2.dev" """, language="toml")
    if has_r2_config():
        st.success("R2 secrets настроены.")
        st.write("Bucket:", st.secrets["R2_BUCKET_NAME"])
        st.write("Public URL:", st.secrets["R2_PUBLIC_BASE_URL"])
        st.write("Кэш проверенных R2-ключей в текущей сессии:", len(st.session_state.r2_exists_cache))
    else:
        st.warning("R2 secrets пока не заданы.")

elif page == "Photo Cache":
    st.title("🗂 Photo Cache")
    count = get_photo_cache_count()
    st.metric("Записей в локальном photo-cache", count)
    st.write("Это SQLite-кэш. После первой удачной загрузки фото в R2 связь товара и фото хранится локально и помогает ускорять повторные выгрузки.")
